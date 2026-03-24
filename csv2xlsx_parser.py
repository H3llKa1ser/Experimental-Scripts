#!/usr/bin/env python3
"""
=============================================================
 CSV to XLSX Converter with Formatting (v2 - Leading Zeros Fix)
 Author: Assistant
 Date: 2026
 Description:
   1. Reads a .csv file
   2. Auto-detects delimiter and encoding
   3. Converts to properly formatted .xlsx
   4. Applies headers, column widths, filters, borders
   5. Preserves leading zeros in all numeric-like strings
=============================================================

Usage:
    # Basic conversion
    python3 csv2xlsx.py -i data.csv

    # Custom output name
    python3 csv2xlsx.py -i data.csv -o report.xlsx

    # Custom delimiter
    python3 csv2xlsx.py -i data.csv -d ";"

    # No formatting (plain)
    python3 csv2xlsx.py -i data.csv --plain

    # Multiple files at once
    python3 csv2xlsx.py -i file1.csv file2.csv file3.csv

    # Merge multiple CSVs into one XLSX (separate sheets)
    python3 csv2xlsx.py -i file1.csv file2.csv --merge

    # Force all columns as text (safest for IDs/codes)
    python3 csv2xlsx.py -i data.csv --force-text

Dependencies:
    pip install openpyxl
"""

import argparse
import csv
import os
import sys
import re
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    print("""
[ERROR] openpyxl is required but not installed.
        Install it with:
            pip install openpyxl
        or:
            pip3 install openpyxl
    """)
    sys.exit(1)


# ============================================================
# CONFIGURATION
# ============================================================

TIMESTAMP = datetime.now().strftime("%Y%m%d_%H%M%S")

# Alternative color themes
THEMES = {
    'blue': {
        'header_bg': '2F5496',
        'header_font': 'FFFFFF',
        'even_bg': 'D6E4F0',
        'odd_bg': 'FFFFFF',
        'border': 'B4C6E7',
    },
    'green': {
        'header_bg': '548235',
        'header_font': 'FFFFFF',
        'even_bg': 'E2EFDA',
        'odd_bg': 'FFFFFF',
        'border': 'A9D18E',
    },
    'dark': {
        'header_bg': '333333',
        'header_font': 'FFFFFF',
        'even_bg': 'F2F2F2',
        'odd_bg': 'FFFFFF',
        'border': 'CCCCCC',
    },
    'red': {
        'header_bg': 'C00000',
        'header_font': 'FFFFFF',
        'even_bg': 'FCE4EC',
        'odd_bg': 'FFFFFF',
        'border': 'E57373',
    },
    'orange': {
        'header_bg': 'ED7D31',
        'header_font': 'FFFFFF',
        'even_bg': 'FBE5D6',
        'odd_bg': 'FFFFFF',
        'border': 'F4B183',
    },
}


# ============================================================
# DELIMITER DETECTION
# ============================================================

def detect_csv_delimiter(filepath, sample_lines=5):
    """
    Auto-detects CSV delimiter by analyzing first N lines.
    Checks: comma, semicolon, tab, pipe
    """
    delimiters = [',', ';', '\t', '|']
    scores = {d: 0 for d in delimiters}

    with open(filepath, 'r', encoding='utf-8', errors='replace') as f:
        lines = []
        for _ in range(sample_lines):
            line = f.readline()
            if line:
                lines.append(line)

    if not lines:
        return ','

    for line in lines:
        for d in delimiters:
            count = line.count(d)
            scores[d] += count

    best = max(scores, key=scores.get)

    if scores[best] == 0:
        return ','

    delim_names = {',': 'comma', ';': 'semicolon', '\t': 'tab', '|': 'pipe'}
    print(f"[INFO] Auto-detected delimiter: {delim_names.get(best, best)}")
    return best


# ============================================================
# ENCODING DETECTION
# ============================================================

def detect_encoding(filepath):
    """Try common encodings."""
    encodings = ['utf-8', 'utf-8-sig', 'latin-1', 'cp1252', 'iso-8859-1']

    for enc in encodings:
        try:
            with open(filepath, 'r', encoding=enc) as f:
                f.read(1024)
            return enc
        except (UnicodeDecodeError, UnicodeError):
            continue

    return 'utf-8'


# ============================================================
# DATA TYPE DETECTION (v2 - LEADING ZEROS FIX)
# ============================================================

def detect_cell_type(value, force_text=False):
    """
    Detects if a cell value is numeric, date, percentage, or text.
    PRESERVES leading zeros by treating them as text.

    Returns: (converted_value, type_string)

    v2 Changes:
        - Leading zero numbers → kept as text
        - Long digit strings → kept as text
        - force_text mode → everything stays as string
    """
    value = str(value).strip()

    if not value:
        return value, 'empty'

    # ---- FORCE TEXT MODE ----
    if force_text:
        return value, 'text_forced'

    # ---- FIX: Leading zeros → ALWAYS treat as text ----
    # Matches: 001234, 00123456789012, 0042, etc.
    if re.match(r'^0\d+$', value):
        return value, 'leading_zero'

    # ---- Long digit strings (10+ digits) → text ----
    # Covers: account IDs, phone numbers, UPC codes, etc.
    if re.match(r'^\d{10,}$', value):
        return value, 'long_number'

    # ---- Zip codes (5 digits starting with 0) ----
    if re.match(r'^0\d{4}$', value):
        return value, 'leading_zero'

    # ---- Regular integer (NO leading zeros, < 10 digits) ----
    if re.match(r'^-?[1-9]\d*$', value) and len(value) < 10:
        return int(value), 'integer'

    # ---- Single zero ----
    if value == '0':
        return 0, 'integer'

    # ---- Float ----
    if re.match(r'^-?\d+\.\d+$', value):
        return float(value), 'float'

    # ---- Percentage ----
    if re.match(r'^-?\d+\.?\d*%$', value):
        return float(value.replace('%', '')) / 100, 'percentage'

    # ---- Currency ----
    if re.match(r'^[$€£]\s?-?\d+[,.]?\d*$', value):
        return value, 'currency'

    # ---- Default: text ----
    return value, 'text'


# ============================================================
# CSV READER
# ============================================================

def read_csv_file(filepath, delimiter=None):
    """
    Reads a CSV file and returns headers + rows.
    """
    if not os.path.exists(filepath):
        print(f"[ERROR] File not found: {filepath}")
        sys.exit(1)

    encoding = detect_encoding(filepath)
    print(f"[INFO] Encoding: {encoding}")

    if delimiter is None:
        delimiter = detect_csv_delimiter(filepath)

    headers = []
    rows = []

    with open(filepath, 'r', encoding=encoding, errors='replace') as f:
        reader = csv.reader(f, delimiter=delimiter, quotechar='"')

        for i, row in enumerate(reader):
            cleaned = [field.strip() for field in row]

            if i == 0:
                headers = cleaned
            else:
                if any(field for field in cleaned):
                    rows.append(cleaned)

    print(f"[OK] Read {len(rows)} rows × {len(headers)} columns")
    return headers, rows


# ============================================================
# THEME BUILDER
# ============================================================

def apply_theme(theme_name='blue'):
    """
    Returns style objects based on selected theme.
    """
    theme = THEMES.get(theme_name, THEMES['blue'])

    header_font = Font(name='Calibri', bold=True, color=theme['header_font'], size=11)
    header_fill = PatternFill(start_color=theme['header_bg'], end_color=theme['header_bg'], fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    even_fill = PatternFill(start_color=theme['even_bg'], end_color=theme['even_bg'], fill_type='solid')
    odd_fill = PatternFill(start_color=theme['odd_bg'], end_color=theme['odd_bg'], fill_type='solid')

    data_font = Font(name='Calibri', size=10)
    data_align = Alignment(horizontal='left', vertical='center')

    border = Border(
        left=Side(style='thin', color=theme['border']),
        right=Side(style='thin', color=theme['border']),
        top=Side(style='thin', color=theme['border']),
        bottom=Side(style='thin', color=theme['border']),
    )

    return {
        'header_font': header_font,
        'header_fill': header_fill,
        'header_align': header_align,
        'even_fill': even_fill,
        'odd_fill': odd_fill,
        'data_font': data_font,
        'data_align': data_align,
        'border': border,
    }


# ============================================================
# CELL WRITER (v2 - SAFE WRITER)
# ============================================================

def write_cell_safe(ws, row_idx, col_idx, value, dtype, styles,
                    is_even, plain):
    """
    Writes a single cell with proper formatting.
    Handles leading zeros and data type formatting safely.

    This is the core fix — ensures leading zeros and long numbers
    are written as TEXT with '@' format.
    """
    cell = ws.cell(row=row_idx, column=col_idx)

    # ---- LEADING ZEROS / LONG NUMBERS / FORCED TEXT ----
    # Must set number_format BEFORE setting value
    if dtype in ('leading_zero', 'long_number', 'text_forced'):
        cell.number_format = '@'        # '@' = Excel text format
        cell.value = str(value)         # Force string

    # ---- PERCENTAGE ----
    elif dtype == 'percentage':
        cell.value = value
        cell.number_format = '0.00%'

    # ---- FLOAT ----
    elif dtype == 'float':
        cell.value = value
        cell.number_format = '#,##0.00'

    # ---- INTEGER ----
    elif dtype == 'integer':
        cell.value = value
        cell.number_format = '#,##0'

    # ---- EVERYTHING ELSE (text, currency, empty) ----
    else:
        cell.value = value

    # ---- APPLY STYLING ----
    if not plain:
        cell.font = styles['data_font']
        cell.fill = styles['even_fill'] if is_even else styles['odd_fill']
        cell.border = styles['border']

        # Right-align numbers
        if dtype in ('integer', 'float', 'percentage'):
            cell.alignment = Alignment(horizontal='right', vertical='center')
        else:
            cell.alignment = styles['data_align']

    return cell


# ============================================================
# XLSX BUILDER (v2 - WITH SAFE CELL WRITING)
# ============================================================

def build_xlsx(headers, rows, output_path, sheet_name="Data",
               theme_name='blue', plain=False, freeze_header=True,
               auto_filter=True, auto_width=True, force_text=False):
    """
    Builds a formatted XLSX file from headers + rows.
    v2: Uses write_cell_safe() to preserve leading zeros.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    styles = apply_theme(theme_name)

    # Track data type statistics
    type_stats = {}

    # ---- WRITE HEADERS ----
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        if not plain:
            cell.font = styles['header_font']
            cell.fill = styles['header_fill']
            cell.alignment = styles['header_align']
            cell.border = styles['border']

    # ---- WRITE DATA ROWS ----
    col_widths = [len(str(h)) + 2 for h in headers]

    for row_idx, row in enumerate(rows, 2):
        is_even = (row_idx % 2 == 0)

        for col_idx, value in enumerate(row, 1):
            if col_idx > len(headers):
                break

            # Detect data type (with leading zero protection)
            converted, dtype = detect_cell_type(value, force_text)

            # Track stats
            type_stats[dtype] = type_stats.get(dtype, 0) + 1

            # Write cell safely
            write_cell_safe(
                ws=ws,
                row_idx=row_idx,
                col_idx=col_idx,
                value=converted,
                dtype=dtype,
                styles=styles,
                is_even=is_even,
                plain=plain
            )

            # Track max column width
            display_len = len(str(value)) + 2
            if col_idx <= len(col_widths):
                col_widths[col_idx - 1] = max(col_widths[col_idx - 1], display_len)

    # ---- AUTO-FIT COLUMN WIDTHS ----
    if auto_width:
        for col_idx, width in enumerate(col_widths, 1):
            adjusted_width = min(width + 3, 50)
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    # ---- FREEZE TOP ROW ----
    if freeze_header:
        ws.freeze_panes = 'A2'

    # ---- AUTO-FILTER ----
    if auto_filter and headers:
        last_col = get_column_letter(len(headers))
        last_row = len(rows) + 1
        ws.auto_filter.ref = f"A1:{last_col}{last_row}"

    # ---- PRINT OPTIONS ----
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    # ---- SAVE ----
    wb.save(output_path)
    print(f"[OK] Saved: {output_path}")

    # Print type detection stats
    print(f"\n[INFO] Data type detection summary:")
    for dtype, count in sorted(type_stats.items(), key=lambda x: -x[1]):
        preserved = " ← PRESERVED" if dtype in ('leading_zero', 'long_number', 'text_forced') else ""
        print(f"       {dtype:<20} {count:>6} cells{preserved}")

    return wb


def build_merged_xlsx(file_list, output_path, delimiter=None,
                      theme_name='blue', plain=False, force_text=False):
    """
    Merges multiple CSV files into one XLSX with separate sheets.
    v2: Uses write_cell_safe() for all data.
    """
    wb = Workbook()
    wb.remove(wb.active)

    for filepath in file_list:
        print(f"\n[INFO] Processing: {filepath}")
        headers, rows = read_csv_file(filepath, delimiter)

        sheet_name = os.path.splitext(os.path.basename(filepath))[0][:31]

        ws = wb.create_sheet(title=sheet_name)
        styles = apply_theme(theme_name)

        # Write headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            if not plain:
                cell.font = styles['header_font']
                cell.fill = styles['header_fill']
                cell.alignment = styles['header_align']
                cell.border = styles['border']

        # Write data
        col_widths = [len(str(h)) + 2 for h in headers]

        for row_idx, row in enumerate(rows, 2):
            is_even = (row_idx % 2 == 0)
            for col_idx, value in enumerate(row, 1):
                if col_idx > len(headers):
                    break

                converted, dtype = detect_cell_type(value, force_text)

                write_cell_safe(
                    ws=ws,
                    row_idx=row_idx,
                    col_idx=col_idx,
                    value=converted,
                    dtype=dtype,
                    styles=styles,
                    is_even=is_even,
                    plain=plain
                )

                display_len = len(str(value)) + 2
                if col_idx <= len(col_widths):
                    col_widths[col_idx - 1] = max(col_widths[col_idx - 1], display_len)

        # Auto-fit widths
        for col_idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = min(width + 3, 50)

        # Freeze & filter
        ws.freeze_panes = 'A2'
        if headers:
            last_col = get_column_letter(len(headers))
            last_row = len(rows) + 1
            ws.auto_filter.ref = f"A1:{last_col}{last_row}"

    wb.save(output_path)
    print(f"\n[OK] Merged {len(file_list)} files → {output_path}")
    return wb


# ============================================================
# REPORT
# ============================================================

def print_report(input_files, output_file, headers, total_rows,
                 theme_name, plain, force_text):
    """Prints conversion summary."""
    report = f"""
╔══════════════════════════════════════════════════════════════╗
║        ✅ CSV → XLSX CONVERSION COMPLETE (v2)               ║
╠══════════════════════════════════════════════════════════════╣
║                                                              ║
║  📥 Input file(s):                                           ║"""

    for f in input_files:
        report += f"\n║     {f:<55}║"

    ft_status = "ON (all text)" if force_text else "OFF (auto-detect)"

    report += f"""
║                                                              ║
║  📊 Output file:   {output_file:<40}║
║                                                              ║
╠══════════════════════════════════════════════════════════════╣
║  DETAILS                                                     ║
╠══════════════════════════════════════════════════════════════╣
║  Total rows:       {total_rows:<40}║
║  Total columns:    {len(headers):<40}║
║  Theme:            {theme_name:<40}║
║  Formatted:        {'No (plain)' if plain else 'Yes':<40}║
║  Frozen header:    {'Yes':<40}║
║  Auto-filter:      {'Yes':<40}║
║  Auto-fit widths:  {'Yes':<40}║
║  Force text mode:  {ft_status:<40}║
║  Leading zeros:    {'PRESERVED ✅':<40}║
║                                                              ║
╠══════════════════════════════════════════════════════════════╣
║  COLUMNS                                                     ║
╠══════════════════════════════════════════════════════════════╣"""

    for i, h in enumerate(headers):
        report += f"\n║  [{i}] {h:<53}║"

    report += """
║                                                              ║
╠══════════════════════════════════════════════════════════════╣
║  LEADING ZEROS FIX (v2)                                      ║
╠══════════════════════════════════════════════════════════════╣
║  ✅ 001234567890 → 001234567890  (preserved)                 ║
║  ✅ 0042         → 0042          (preserved)                 ║
║  ✅ 00001        → 00001         (preserved)                 ║
║  ✅ 123456789012 → 123456789012  (12-digit, kept as text)    ║
║  ✅ 42           → 42            (normal integer)            ║
║  ✅ 3.14         → 3.14          (normal float)              ║
║                                                              ║
╚══════════════════════════════════════════════════════════════╝
"""
    print(report)


# ============================================================
# MAIN
# ============================================================

def main():
    parser = argparse.ArgumentParser(
        description="📊 CSV to XLSX Converter v2 (Leading Zeros Fix)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Basic conversion
  python3 csv2xlsx.py -i data.csv

  # Custom output filename
  python3 csv2xlsx.py -i data.csv -o report.xlsx

  # Semicolon-delimited CSV
  python3 csv2xlsx.py -i data.csv -d ";"

  # Green theme
  python3 csv2xlsx.py -i data.csv --theme green

  # Plain (no formatting)
  python3 csv2xlsx.py -i data.csv --plain

  # Force ALL columns as text (safest for IDs/codes)
  python3 csv2xlsx.py -i data.csv --force-text

  # Multiple files → separate XLSX each
  python3 csv2xlsx.py -i file1.csv file2.csv file3.csv

  # Multiple files → one XLSX with multiple sheets
  python3 csv2xlsx.py -i file1.csv file2.csv --merge

  # Custom sheet name
  python3 csv2xlsx.py -i data.csv --sheet "Health Check"

  # Full workflow with masker script
  python3 txt2csv_mask.py -i data.txt --group "3:4" --headers "ID,Resource,Type,Access,Name,Status"
  python3 csv2xlsx.py -i data_masked_*.csv -o report.xlsx --theme blue

Available Themes:
  blue   - Professional blue (default)
  green  - Nature green
  dark   - Dark headers
  red    - Alert red
  orange - Warm orange

Leading Zeros Handling (v2):
  001234567890 → 001234567890  ✅ Preserved as text
  0042         → 0042          ✅ Preserved as text
  123456789012 → 123456789012  ✅ Long number, kept as text
  42           → 42            ✅ Normal integer
  3.14         → 3.14          ✅ Normal float
  75%          → 0.75          ✅ Percentage format
        """)

    parser.add_argument('-i', '--input', nargs='+', required=True,
                        help='Input CSV file(s)')
    parser.add_argument('-o', '--output', default=None,
                        help='Output XLSX file path')
    parser.add_argument('-d', '--delimiter', default=None,
                        help='CSV delimiter (default: auto-detect)')
    parser.add_argument('--sheet', default='Data',
                        help='Sheet name (default: Data)')
    parser.add_argument('--theme', default='blue',
                        choices=['blue', 'green', 'dark', 'red', 'orange'],
                        help='Color theme (default: blue)')
    parser.add_argument('--plain', action='store_true',
                        help='No formatting (plain conversion)')
    parser.add_argument('--merge', action='store_true',
                        help='Merge multiple CSVs into one XLSX (separate sheets)')
    parser.add_argument('--no-freeze', action='store_true',
                        help='Do not freeze the header row')
    parser.add_argument('--no-filter', action='store_true',
                        help='Do not add auto-filter')
    parser.add_argument('--force-text', action='store_true',
                        help='Force ALL cells as text (safest for IDs/codes)')

    args = parser.parse_args()

    # ---- MERGE MODE ----
    if args.merge and len(args.input) > 1:
        output = args.output or f"merged_{TIMESTAMP}.xlsx"
        build_merged_xlsx(
            file_list=args.input,
            output_path=output,
            delimiter=args.delimiter,
            theme_name=args.theme,
            plain=args.plain,
            force_text=args.force_text
        )
        print(f"\n✅ Done! Merged {len(args.input)} files → {output}")
        return

    # ---- SINGLE / MULTI FILE MODE ----
    for input_file in args.input:
        print(f"\n{'='*60}")
        print(f"[INFO] Processing: {input_file}")
        print(f"{'='*60}")

        # Read CSV
        headers, rows = read_csv_file(input_file, args.delimiter)

        # Generate output filename
        if args.output and len(args.input) == 1:
            output_file = args.output
        else:
            base = os.path.splitext(os.path.basename(input_file))[0]
            output_file = f"{base}_{TIMESTAMP}.xlsx"

        # Build XLSX
        build_xlsx(
            headers=headers,
            rows=rows,
            output_path=output_file,
            sheet_name=args.sheet,
            theme_name=args.theme,
            plain=args.plain,
            freeze_header=not args.no_freeze,
            auto_filter=not args.no_filter,
            force_text=args.force_text,
        )

        # Report
        print_report(
            input_files=[input_file],
            output_file=output_file,
            headers=headers,
            total_rows=len(rows),
            theme_name=args.theme,
            plain=args.plain,
            force_text=args.force_text,
        )

    print(f"\n✅ All done! Processed {len(args.input)} file(s).")


if __name__ == "__main__":
    main()
