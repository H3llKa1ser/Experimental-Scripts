import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import DataBarRule
import os
import sys


def beautify_excel(input_file, output_file=None):
    """
    Takes an Excel file and transforms its tables into beautified,
    actionable, sortable, and easily modifiable formats.
    """
    if output_file is None:
        name, ext = os.path.splitext(input_file)
        output_file = f"{name}_beautified{ext}"

    wb = openpyxl.load_workbook(input_file)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n📄 Processing sheet: '{sheet_name}'")

        # Skip empty sheets
        if ws.max_row is None or ws.max_row < 2 or ws.max_column is None:
            print("   ⚠️  Sheet is empty or has no data rows. Skipping.")
            continue

        max_row = ws.max_row
        max_col = ws.max_column

        # ─────────────────────────────────────────────
        # 1. DEFINE STYLES
        # ─────────────────────────────────────────────
        header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        data_font = Font(name="Calibri", size=11)
        data_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        number_alignment = Alignment(horizontal="right", vertical="center")

        thin_border = Border(
            left=Side(style="thin", color="B4C6E7"),
            right=Side(style="thin", color="B4C6E7"),
            top=Side(style="thin", color="B4C6E7"),
            bottom=Side(style="thin", color="B4C6E7"),
        )

        even_row_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
        odd_row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        # ─────────────────────────────────────────────
        # 2. STYLE THE HEADER ROW
        # ─────────────────────────────────────────────
        for col in range(1, max_col + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # ─────────────────────────────────────────────
        # 3. STYLE DATA ROWS (Alternating Colors)
        # ─────────────────────────────────────────────
        for row in range(2, max_row + 1):
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = data_font
                cell.border = thin_border

                # Alternating row colors
                if row % 2 == 0:
                    cell.fill = even_row_fill
                else:
                    cell.fill = odd_row_fill

                # Align numbers to the right, text to the left
                if isinstance(cell.value, (int, float)):
                    cell.alignment = number_alignment
                    # Add thousand separator formatting for large numbers
                    if isinstance(cell.value, float):
                        cell.number_format = '#,##0.00'
                    elif isinstance(cell.value, int) and abs(cell.value) >= 1000:
                        cell.number_format = '#,##0'
                else:
                    cell.alignment = data_alignment

        # ─────────────────────────────────────────────
        # 4. AUTO-FIT COLUMN WIDTHS
        # ─────────────────────────────────────────────
        for col in range(1, max_col + 1):
            max_length = 0
            col_letter = get_column_letter(col)
            for row in range(1, max_row + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value is not None:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            # Add padding, cap at 50
            adjusted_width = min(max_length + 4, 50)
            ws.column_dimensions[col_letter].width = max(adjusted_width, 12)

        # ─────────────────────────────────────────────
        # 5. ADD EXCEL TABLE WITH SORT & FILTER
        # ─────────────────────────────────────────────
        table_ref = f"A1:{get_column_letter(max_col)}{max_row}"
        table_name = f"Table_{sheet_name.replace(' ', '_')}"

        # Remove any existing tables to avoid conflicts
        if ws.tables:
            for t in list(ws.tables.keys()):
                del ws.tables[t]

        table = Table(displayName=table_name, ref=table_ref)
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style
        ws.add_table(table)
        print(f"   ✅ Added sortable/filterable table: {table_name} ({table_ref})")

        # ─────────────────────────────────────────────
        # 6. ADD DATA BARS FOR NUMERIC COLUMNS
        # ─────────────────────────────────────────────
        for col in range(1, max_col + 1):
            # Check if the column is numeric (sample first few data rows)
            is_numeric = True
            for row in range(2, min(max_row + 1, 12)):  # Check up to 10 rows
                val = ws.cell(row=row, column=col).value
                if val is not None and not isinstance(val, (int, float)):
                    is_numeric = False
                    break

            if is_numeric and max_row > 1:
                col_letter = get_column_letter(col)
                data_range = f"{col_letter}2:{col_letter}{max_row}"
                rule = DataBarRule(
                    start_type="min",
                    end_type="max",
                    color="5B9BD5",
                    showValue=True,
                    minLength=None,
                    maxLength=None,
                )
                ws.conditional_formatting.add(data_range, rule)
                print(f"   📊 Added data bars to column '{ws.cell(1, col).value}' ({col_letter})")

        # ─────────────────────────────────────────────
        # 7. FREEZE THE HEADER ROW
        # ─────────────────────────────────────────────
        ws.freeze_panes = "A2"
        print("   ❄️  Froze header row for easy scrolling")

        # ─────────────────────────────────────────────
        # 8. SET PRINT AREA & PAGE SETUP
        # ─────────────────────────────────────────────
        ws.print_area = table_ref
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.print_title_rows = "1:1"  # Repeat header on every printed page
        print("   🖨️  Set print area & landscape orientation")

        # ─────────────────────────────────────────────
        # 9. ADD ROW HEIGHT FOR READABILITY
        # ─────────────────────────────────────────────
        ws.row_dimensions[1].height = 30  # Header row height
        for row in range(2, max_row + 1):
            ws.row_dimensions[row].height = 22

    # ─────────────────────────────────────────────
    # SAVE THE BEAUTIFIED FILE
    # ─────────────────────────────────────────────
    wb.save(output_file)
    print(f"\n🎉 Beautified file saved to: {output_file}")
    return output_file


# ═══════════════════════════════════════════════════
# MAIN EXECUTION
# ═══════════════════════════════════════════════════
if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python beautify_excel.py <input_file.xlsx> [output_file.xlsx]")
        print("Example: python beautify_excel.py sales_data.xlsx sales_data_pretty.xlsx")
        sys.exit(1)

    input_path = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else None

    if not os.path.exists(input_path):
        print(f"❌ Error: File '{input_path}' not found.")
        sys.exit(1)

    beautify_excel(input_path, output_path)
